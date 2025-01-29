from django.contrib import messages
from django.shortcuts import render, redirect
from django.core.exceptions import ValidationError
from .forms import KirdiChiqdiUploadForm
from .models import Mahsulot, OlchovBirligi, KirdiChiqdi
from datetime import datetime
import openpyxl
from decimal import Decimal


def kirdi_upload_view(request):
    # Hozirgi vaqtni olish
    now = datetime.now()

    if request.method == "POST":
        form = KirdiChiqdiUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data["file"]
            try:
                # Excel faylni o'qiymiz
                wb = openpyxl.load_workbook(file)
                sheet = wb.active

                # Fayldagi validatsiya
                required_columns = 4  # kerakli ustunlar soni (endilikda summa ustuni qo'shildi)
                if sheet.max_column < required_columns:
                    messages.error(request, "Faylda kerakli ustunlar yetarli emas!")
                    return redirect("admin:index")

                # Fayldagi ma'lumotlarni qayta ishlash
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Birinchi qatordan tashqari
                    if len(row) < required_columns or not all(row[:required_columns]):
                        messages.warning(request, f"Ma'lumot to'liq emas: {row}")
                        continue

                    mahsulot_nomi = row[0].strip()
                    miqdor = row[1]
                    olchov_birligi_nomi = row[2].strip()
                    summa = row[3]  # Yangi summa ustuni

                    # Summa tekshiruvi
                    try:
                        summa = Decimal(summa)  # Summa decimalga o'tkaziladi
                        if summa <= 0:
                            messages.warning(request, f"Summa musbat bo'lishi kerak: {row}")
                            continue
                    except (ValueError, TypeError):
                        messages.warning(request, f"Summa noto'g'ri formatda: {row}")
                        continue

                    # O'lchov birligi bazada bormi yoki yaratamiz
                    olchov_birligi, _ = OlchovBirligi.objects.get_or_create(
                        olchov_birligi=olchov_birligi_nomi.title()
                    )

                    # Mahsulotni yaratish yoki olish
                    try:
                        mahsulot = Mahsulot.objects.filter(
                            mahsulot_nomi__iexact=mahsulot_nomi.strip(),
                            olchov_birligi=olchov_birligi
                        ).first()

                        if mahsulot is None:
                            mahsulot = Mahsulot(
                                mahsulot_nomi=mahsulot_nomi,
                                olchov_birligi=olchov_birligi
                            )
                            mahsulot.clean()  # Validatsiya qilish
                            mahsulot.save()  # Saqlash
                    except ValidationError as e:
                        messages.warning(request, f"Xatolik: {e.message_dict.get('mahsulot_nomi', 'Xato!')}")
                        continue

                    # KirdiChiqdi modeliga ma'lumotni saqlash
                    KirdiChiqdi.objects.create(
                        mahsulot_nomi=mahsulot,
                        miqdor=miqdor,
                        summa=summa,  # Summa qo'shildi
                        sana=now,
                        amaliyot_turi="Kirdi"
                    )
                messages.success(request, "Fayl muvaffaqiyatli yuklandi va ma'lumotlar saqlandi!")
                return redirect("admin:index")
            except Exception as e:
                messages.error(request, f"Xatolik yuz berdi: {e}")
    else:
        form = KirdiChiqdiUploadForm()

    return render(request, "admin/kirdi_upload.html", {"form": form})




# from django.contrib import messages
# from django.shortcuts import render, redirect
# from django.core.exceptions import ValidationError
# from .forms import KirdiChiqdiUploadForm
# from .models import Mahsulot, OlchovBirligi, KirdiChiqdi
# from datetime import datetime
# import openpyxl
#
#
# def kirdi_upload_view(request):
#     # Hozirgi vaqtni olish
#     now = datetime.now()
#
#     if request.method == "POST":
#         form = KirdiChiqdiUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             file = form.cleaned_data["file"]
#             try:
#                 # Excel faylni o'qiymiz
#                 wb = openpyxl.load_workbook(file)
#                 sheet = wb.active
#
#                 # Fayldagi validatsiya
#                 required_columns = 3  # kerakli ustunlar soni
#                 if sheet.max_column < required_columns:
#                     messages.error(request, "Faylda kerakli ustunlar yetarli emas!")
#                     return redirect("admin:index")
#
#                 # Fayldagi ma'lumotlarni qayta ishlash
#                 for row in sheet.iter_rows(min_row=2, values_only=True):  # Birinchi qatordan tashqari
#                     if len(row) < required_columns or not all(row[:required_columns]):
#                         messages.warning(request, f"Ma'lumot to'liq emas: {row}")
#                         continue
#
#                     mahsulot_nomi = row[0].strip()
#                     miqdor = row[1]
#                     olchov_birligi_nomi = row[2].strip()
#
#                     # O'lchov birligi bazada bormi yoki yaratamiz
#                     olchov_birligi, _ = OlchovBirligi.objects.get_or_create(
#                         olchov_birligi=olchov_birligi_nomi.title()
#                     )
#
#                     # Mahsulotni yaratish yoki olish
#                     try:
#                         mahsulot = Mahsulot.objects.filter(
#                             mahsulot_nomi__iexact=mahsulot_nomi.strip(),
#                             olchov_birligi=olchov_birligi
#                         ).first()
#
#                         if mahsulot is None:
#                             mahsulot = Mahsulot(
#                                 mahsulot_nomi=mahsulot_nomi,
#                                 olchov_birligi=olchov_birligi
#                             )
#                             mahsulot.clean()  # Validatsiya qilish
#                             mahsulot.save()  # Saqlash
#                     except ValidationError as e:
#                         messages.warning(request, f"Xatolik: {e.message_dict.get('mahsulot_nomi', 'Xato!')}")
#                         continue
#
#                     # KirdiChiqdi modeliga ma'lumotni saqlash
#                     KirdiChiqdi.objects.create(
#                         mahsulot_nomi=mahsulot,
#                         miqdor=miqdor,
#                         sana=now,
#                         amaliyot_turi="Kirdi"
#                     )
#                 messages.success(request, "Fayl muvaffaqiyatli yuklandi va ma'lumotlar saqlandi!")
#                 return redirect("admin:index")
#             except Exception as e:
#                 messages.error(request, f"Xatolik yuz berdi: {e}")
#     else:
#         form = KirdiChiqdiUploadForm()
#
#     return render(request, "admin/kirdi_upload.html", {"form": form})
